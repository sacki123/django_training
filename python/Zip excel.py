import os
import datetime
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.cell.cell import TYPE_NUMERIC
from openpyxl.styles.numbers import FORMAT_TEXT
from openpyxl.utils.cell import get_column_letter

list_file= os.listdir(os.path.join(os.path.dirname(__file__), "input"))
input_path = os.path.join(os.path.dirname(__file__), "input")
class main_class():
    __h = "hoang"
    h1 = "123"

    def __init__(self):
        self.name = "ホアン"
        self.__h = "ABC"

    def read_sheet_name(self,path_file):
        wb = load_workbook(path_file, data_only=True)
        return wb

    def create_list(self):
        excel_group = []
        for excel_file in list_file:
            wb = self.read_sheet_name(os.path.join(input_path,excel_file))
            name_sheet = wb.sheetnames
            for j in name_sheet:   
                ws = wb[j]
                rows = ws.max_row
                for value in ws.values:
                    excel_group.append(list(value))
                    
        return excel_group

    def set_auto_col_width(self,sheet):
        dimensions = {}
        for row in sheet.rows:
            for cell in row:
                if cell.value:
                    width = max((dimensions.get(cell.column_letter, 0), len(str(cell.value))))
                    if width > 40:
                        width = 40
                    elif width < 10:
                        width = 10    
                    dimensions[cell.column_letter] = width
        for col, value in dimensions.items():
            sheet.column_dimensions[col].width = value * 2 + (3 * (value / 6.5))

    def format_number(self,cell):
        cell.number_format = '#,##0'

    def format_text(self,cell):
        cell.number_format = FORMAT_TEXT

    def set_wrap_text(self,cell):
        cell.alignment = Alignment(wrapText=True)

    def sheet_format_number(self,sheet, no_format_columns=None):
        if not no_format_columns:
            no_format_columns = []

        for row in sheet.rows:
            for cell in row:
                if cell.column_letter not in no_format_columns and cell.data_type == TYPE_NUMERIC:
                    self.format_number(cell)
                elif cell.is_date:
                    pass
                else:
                    self.format_text(cell)
                
                if '\n' in str(cell.value):
                    self.set_wrap_text(cell)

    def create_excel(self,dic_main):
        wb = Workbook()
        ws = wb.active
        ws.title = "エラー"
        time = datetime.datetime.now().strftime("%y%m%d%H%M%S")
        excel_name = time + "エラー一覧まとめ.xlsx"
        excel_path = os.path.join(os.path.dirname(__file__),excel_name ) 
        for row in dic_main:
            ws.append(row)
        self.set_auto_col_width(ws) 
        self.sheet_format_number(ws)   
        wb.save(excel_path)    
    
if (__name__ == "__main__" ):   
    new = main_class()         
    main_list = new.create_list()
    new.create_excel(main_list)
print("OK")    
