#!/Commonusr/bin/python
# -*- coding: utf-8 -*-
"""
帳票(PDF) パッケージ


"""

__title__ = '帳票(PDF) パッケージ'
__author__ = "ExpertSoftware Inc."
__status__ = "develop"
__version__ = "0.0.0_0"
__date__ = "2018/07/27"
__license__ = ''
__desc__ = '%s Ver%s (%s)' % (__title__, __version__, __date__)

# ------------------------------------------------------------
# Import Section
# ------------------------------------------------------------
# Python
import io
import copy
import langid
import traceback
import itertools
from logging import getLogger

# PyPDF
from PyPDF2.pdf import PdfFileReader
from PIL import ImageFont

# reportlab
from reportlab.pdfgen import canvas
from reportlab.pdfbase import pdfmetrics
from reportlab.lib.pagesizes import letter
from reportlab.pdfbase.ttfonts import TTFont

# Excel
import xlwt
from xlrd import open_workbook
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
from xlsxwriter.utility import xl_rowcol_to_cell

# zen
from zen.pdf import AbstractCommand
from zen.db.mongo import ZenMongoDB

# constant
from zen.pdf import constant

# s3
from zen.s3 import set_resource

# ------------------------------------------------------------
# Variable Section
# ------------------------------------------------------------
log = getLogger(__name__)
# ------------------------------------------------------------
# Class Section
# ------------------------------------------------------------


class MakeCommand(AbstractCommand):
    """
    PDF作成＋格納クラス
    """

    def __init__(self, config, unit, command="Make"):
        """
        コンストラクタ
        """
        super().__init__(config, unit, command)

    def onExecute(self):
        """
        処理実行
        @return 終了コード
        """
        # ログ
        log.debug(self)
        # メイン処理
        exit_code = self.func.execute(self)
        # 問題無ければPDFをテンプレートを合成し保存
        if exit_code > 0:
            pass
        return exit_code

    def onPrepare(self):
        """
        処理前準備
        @return 終了コード
        """
        # ログ
        log.debug(self)
        try:
            # メモリ処理用
            self.files = FilesMaker(self.unit, self.category, self.kinoubango, self.activity_code, self.config)
        except:
            # 例外処理
            log.error(traceback.format_exc())
        # ログ
        log.info(self)
        return 1


class CommonFilesMaker(object):

    def __init__(self, module, category, kinoubango, activity_code, config):
        self.config = config
        self.module = module
        self.activity_code = activity_code
        self.category = category
        self.kinoubango = kinoubango
        self.mongodb = ZenMongoDB()
        self.file_output_name = constant.set_output_file_name_by_kinoubango(self.kinoubango, self.category,
                                                                            self.activity_code)
        self.json_data_list = self.mongodb.find_data(self.module, self.mongodb.collection)
        if self.json_data_list is not None:
            for i in self.json_data_list:
                if self.kinoubango in i.get("REPORT_NUMBER"):
                    self.json_data = i.get("REPORT_DATA")
        self.json_data = self.json_data if self.json_data_list else constant.read_json(constant.get_json_file_by_kinoubango(self.kinoubango, self.category, self.activity_code))
        self.s3_info = self.set_info_s3(self.get_config_on_json())
        try:
            if self.s3_info.get("use_ssl") == 'False':
                use_ssl = False
            elif self.s3_info.get("use_ssl") == 'True':
                use_ssl = True
        except:
            constant.get_error(constant.ERROR_004)
        try:
            set_resource(
                self.s3_info.get("profile_name"),
                self.s3_info.get("aws_access_key_id"),
                self.s3_info.get("aws_secret_access_key"),
                use_ssl,
                self.s3_info.get("endpoint_url")
            )
        except:
            constant.get_error(constant.ERROR_004)
        try:
            self.template = constant.get_template_file_by_kinoubango(self.kinoubango, self.category, self.activity_code,
                                                                 self.s3_info.get("profile_name"))
        except:
            constant.get_error(constant.ERROR_004)

    def set_info_s3(self, config_on_json):
        """
        set information for AWS
        :param config_on_json: config data on json file
        :return: config data for AWS
        """
        try:
            if config_on_json:
                if config_on_json.get("aws_access_key_id") and config_on_json.get("aws_secret_access_key") and config_on_json.get("use_ssl") and config_on_json.get("endpoint_url"):
                    config_on_json["profile_name"] = "json_config"
                    return config_on_json
                else:
                    config_default_dict = {
                        "profile_name": "default",
                        "aws_access_key_id": self.config.get('s3', 'aws_access_key_id'),
                        "aws_secret_access_key": self.config.get('s3', 'aws_secret_access_key'),
                        "use_ssl": self.config.get('s3', 'use_ssl'),
                        "endpoint_url": self.config.get('s3', 'endpoint_url'),
                    }
                    return config_default_dict
        except:
            constant.get_error(constant.ERROR_004)
        
    @staticmethod
    def set_spacing(node, number, section_number):
        if number % section_number == 0:
            return [constant.DEFAULT_LINE_SPACING, 0]
        if node.get("SECTION_SPACE") and node.get("LINE_SPACE") and \
                isinstance(node.get("SECTION_SPACE"), dict) and isinstance(node.get("LINE_SPACE"), dict):
            if len(node.get("SECTION_SPACE").items()) == len(node.get("LINE_SPACE").items()):
                line_space = CommonFilesMaker.option_space(number % section_number, node, "LINE_SPACE")
                section_space = CommonFilesMaker.option_space(number % section_number, node, "SECTION_SPACE")
                return [line_space, section_space]
            else:
                return [constant.DEFAULT_LINE_SPACING, node.get("SECTION_SPACE", constant.DEFAULT_SECTION_SPACING)]
        else:
            return [constant.DEFAULT_LINE_SPACING, constant.DEFAULT_SECTION_SPACING]

    @staticmethod
    def option_space(number, node, key):
        result = 0
        if key == "LINE_SPACE":
            result = constant.DEFAULT_LINE_SPACING
        elif key == "SECTION_SPACE":
            result = node.get("SECTION_SPACE", constant.DEFAULT_SECTION_SPACING)
        if node.get(key) and isinstance(node.get(key), dict):
            for index in range(len(node.get(key).values())):
                if index == number - 1:
                    result = list(node.get(key).values())[index]
                    return result
        else:
            return result
        return result

    @staticmethod
    def set_up_loop_page(mapping, loop_mapping, temp, header_info):
        mapping_copy = copy.deepcopy(mapping)
        loop_mapping_copy = copy.deepcopy(loop_mapping)
        header_info_copy = copy.deepcopy(header_info)
        temp.append(mapping_copy[-1])
        for lup_map in loop_mapping_copy[-1]:
            if any(lup_map[0]) and any(lup_map[1]):
                data = list(itertools.chain.from_iterable([header_info_copy[-1][0], lup_map[0]]))
                pos = list(itertools.chain.from_iterable([header_info_copy[-1][1], lup_map[1]]))
                page = lup_map[2]
                temp.append([data, pos, page])

    def get_config_on_json(self):
        """
        Reading CONFIG section on json
        :return: Config content of function
        """
        # load section CONFIG from data
        try:
            return self.json_data["CONFIG"]
        except:
            constant.get_error(constant.ERROR_004)

    def get_data_on_json(self, pos_list):
        """
        Reading ITEM SECTION on json then Mapping data from json to position for each copy document (pdf or excel)
        created
        :param pos_list: position list
        :return: mapping list
        """
        try:
            mapping_list = list()
            for out_put in self.json_data["ITEM"]:
                extra_mapping = []
                header_info = []
                mapping = []
                index = 0
                page = 0
                temp = list()
                for node in out_put:
                    extra_page_number = 0
                    if int(node["PAGE"]) is not page:
                        if index != 0:
                            CommonFilesMaker.set_up_loop_page(mapping, extra_mapping, temp, header_info)
                        index = 0
                        page = int(node["PAGE"])
                        # [0]: json data
                        # [1]: position data
                        # [2]: page number in template
                        # extra_mapping.append([[], [], page])
                        header_info.append([[], [], page])
                        mapping.append([[], [], page])
                        extra_mapping.append([[[], [], page]])
                    self.mapping_position_data(node, mapping, pos_list, index, extra_mapping, extra_page_number,
                                            header_info)
                    index += 1
                    if node is out_put[-1]:
                        CommonFilesMaker.set_up_loop_page(mapping, extra_mapping, temp, header_info)
                mapping_list.append(temp)
                # mapping_list.append(extra_mapping)
            return mapping_list
        except:
            log.error(traceback.format_exc())

    def mapping_position_data(self, node, ol, pl, index, extra_mapping, increase_page_num, header_info):
        """
        Reading ITEM SECTION on json then Mapping data from json to position for each copy document (pdf or excel)
        created
        :param header_info: header of page
        :param node: node of json
        :param ol: output list
        :param pl: position list
        :param index: Number of Node
        :param extra_mapping: data loop pages
        :param increase_page_num:
        """

        page = int(node["PAGE"])
        current_position = pl[page - 1][index][node.get("LOOP_COUNT")] if (
                node.get("LOOP_COUNT", False) is not False) else pl[page - 1][index]
        if node["TYPE"] == "DATA":
            value_list_cnter = 0
            number_loop_element = 0
            for value_list in node["VALUE"]:
                s = 0
                max_section = current_position[s].get("MAX_SECTION", constant.MAX_SECTION_DEFAULT)
                increase_page_num = int(node.get("ITEM_NUMBER", 0) / max_section)
                flag_increase_page = 0
                extra_page_flag_in_data = 0
                if node.get("LOOP_COUNT") is not None:
                    if len(value_list) <= 1:
                        if increase_page_num >= 1:
                            node["INCREASE_PAGE"] = increase_page_num
                            for value in value_list:
                                tem_pos_element = []
                                line_space = CommonFilesMaker.set_spacing(current_position[s], node["SECTION_REPEAT"], max_section)[0]
                                section_space = CommonFilesMaker.set_spacing(current_position[s], node["SECTION_REPEAT"], max_section)[1]
                                for pos_element in current_position[s]["POSITION"]:
                                    tem_pos_element.append([pos_element[0], pos_element[1]
                                                            + node["SECTION_REPEAT"] * section_space
                                                            + number_loop_element * line_space])
                                tem_dict = current_position[s].copy()
                                tem_dict["POSITION"] = tem_pos_element
                                if node["INCREASE_PAGE"] is not flag_increase_page:
                                    extra_mapping[page - 1].append([[], [], page])

                                extra_mapping[page - 1][increase_page_num][0].append(value)
                                extra_mapping[page - 1][increase_page_num][1].append(current_position[s])
                                s += 1
                                flag_increase_page = increase_page_num
                        else:
                            for value in value_list:
                                ol[page - 1][0].append(value)
                                ol[page - 1][1].append(current_position[s])
                                s += 1
                    else:
                        if node.get("FLAG"):
                            if increase_page_num >= 1:
                                node["INCREASE_PAGE"] = increase_page_num
                                for value in value_list:
                                    tem_pos_element = []
                                    line_space = CommonFilesMaker.set_spacing(current_position[s], node["SECTION_REPEAT"], max_section)[0]
                                    section_space = CommonFilesMaker.set_spacing(current_position[s], node["SECTION_REPEAT"], max_section)[1]
                                    for pos_element in current_position[s]["POSITION"]:
                                        tem_pos_element.append([pos_element[0], pos_element[1]
                                                                - (node["SECTION_REPEAT"] % max_section) * section_space
                                                                - (number_loop_element % max_section) * line_space])
                                    tem_dict = current_position[s].copy()
                                    tem_dict["POSITION"] = tem_pos_element
                                    if node["INCREASE_PAGE"] is not flag_increase_page:
                                        extra_mapping[page - 1].append([[], [], page])

                                    extra_mapping[page - 1][increase_page_num][0].append(value)
                                    extra_mapping[page - 1][increase_page_num][1].append(tem_dict)
                                    s += 1
                                number_loop_element += 1
                                flag_increase_page = increase_page_num
                            else:
                                for value in value_list:
                                    tem_pos_element = []
                                    line_space = CommonFilesMaker.set_spacing(current_position[s], node["SECTION_REPEAT"], max_section)[0]
                                    section_space = CommonFilesMaker.set_spacing(current_position[s], node["SECTION_REPEAT"], max_section)[1]
                                    for pos_element in current_position[s]["POSITION"]:
                                        tem_pos_element.append([pos_element[0], pos_element[1]
                                                                - node["SECTION_REPEAT"] * section_space
                                                                - (number_loop_element % max_section) * line_space])
                                    tem_dict = current_position[s].copy()
                                    tem_dict["POSITION"] = tem_pos_element
                                    ol[page - 1][0].append(value)
                                    ol[page - 1][1].append(tem_dict)
                                    s += 1
                                number_loop_element += 1
                        else:
                            for value in value_list:
                                tem_pos_element = []
                                line_space = current_position[s].get("FIRST_LINE_SPACE", constant.DEFAULT_LINE_SPACING)
                                line_space_counter = number_loop_element * int(line_space)
                                for pos_element in current_position[s]["POSITION"]:
                                    tem_pos_element.append([pos_element[0], pos_element[1] - line_space_counter])

                                tem_dict = current_position[s].copy()
                                tem_dict["POSITION"] = tem_pos_element
                                ol[page - 1][0].append(value)
                                ol[page - 1][1].append(tem_dict)
                                s += 1
                            number_loop_element += 1
                else:
                    if increase_page_num >= 1:
                        node["INCREASE_PAGE"] = increase_page_num
                        for value in value_list:
                            if node["INCREASE_PAGE"] is not flag_increase_page:
                                extra_mapping[page - 1].append([[], [], page])
                            extra_mapping[page - 1][increase_page_num][0].append(value)
                            extra_mapping[page - 1][increase_page_num][1].append(current_position[s])
                            s += 1
                            flag_increase_page = increase_page_num
                    else:
                        max_section_in_data_loop = current_position[s].get("MAX_SECTION", constant.MAX_SECTION_DEFAULT)
                        extra_page_number_in_loop_data = int(value_list_cnter / max_section_in_data_loop)
                        if extra_page_number_in_loop_data >= 1:
                            x_flag = extra_page_flag_in_data
                            for value in value_list:
                                tem_pos_element_x = []
                                line_space = current_position[s].get("FIRST_LINE_SPACE", constant.DEFAULT_LINE_SPACING)
                                line_space_counter = (value_list_cnter % max_section_in_data_loop) * int(line_space)
                                for pos_element in current_position[s]["POSITION"]:
                                    tem_pos_element_x.append([pos_element[0], pos_element[1] - line_space_counter])
                                tem_dict = current_position[s].copy()
                                tem_dict["POSITION"] = tem_pos_element_x
                                if x_flag != extra_page_flag_in_data:
                                    extra_mapping[page - 1].append([[], [], page])
                                extra_mapping[page - 1][extra_page_number_in_loop_data - 1][0].append(value)
                                extra_mapping[page - 1][extra_page_number_in_loop_data - 1][1].append(tem_dict)
                                s += 1
                                extra_page_flag_in_data = extra_page_number_in_loop_data
                            value_list_cnter += 1
                        else:
                            for value in value_list:
                                tem_pos_element = []
                                line_space = current_position[s].get("FIRST_LINE_SPACE", constant.DEFAULT_LINE_SPACING)
                                line_space_counter = value_list_cnter * int(line_space)
                                for pos_element in current_position[s]["POSITION"]:
                                    tem_pos_element.append([pos_element[0], pos_element[1] - line_space_counter])
                                tem_dict = current_position[s].copy()
                                tem_dict["POSITION"] = tem_pos_element
                                ol[page - 1][0].append(value)
                                ol[page - 1][1].append(tem_dict)
                                s += 1
                            value_list_cnter += 1
        elif node["TYPE"] == "LINE":
            for value_list in node["VALUE"]:
                s = 0
                for value in value_list:
                    ol[page - 1][0].append(value)
                    ol[page - 1][1].append(current_position[s])
                    s += 1
        elif node["TYPE"] == "HEADER":
            for value_list in node["VALUE"]:
                s = 0
                for value in value_list:
                    ol[page - 1][0].append(value)
                    ol[page - 1][1].append(current_position[s])
                    header_info[page - 1][0].append(value)
                    header_info[page - 1][1].append(current_position[s])
                    s += 1
        elif node["TYPE"] == "SECTION_LOOP":
            cnt_section_loop = 1
            item_number = 0
            for child_nodes in node["ITEM"]:
                i = 0
                if node.get("ITEM"):
                    for node in child_nodes:
                        node["PAGE"] = page
                        node["LOOP_COUNT"] = i
                        node["ITEM_NUMBER"] = item_number
                        i += 1
                        self.mapping_position_data(node, ol, pl, index, extra_mapping, increase_page_num, header_info)
                else:
                    for node in child_nodes:
                        node["PAGE"] = page
                        node["LOOP_COUNT"] = i
                        node["FLAG"] = True
                        node["SECTION_REPEAT"] = cnt_section_loop
                        node["ITEM_NUMBER"] = item_number
                        i += 1
                        self.mapping_position_data(node, ol, pl, index, extra_mapping, increase_page_num, header_info)
                    cnt_section_loop += 1
                item_number += 1


class PDFFilesMaker(CommonFilesMaker):

    def __init__(self, module, category, kinoubango, activity_code, config):
        super().__init__(module, category, kinoubango, activity_code, config)

    # Pages of PDF File
    def create_page_ZU_P99_JA(self, template, page_number, lst_position, lst_data):
        """
        Read context page from file template. Adding specific data to page then return page context
        :param template : File pdf template
        :param page_number: Page Number
        :param lst_position : List positions of each data to adding
        :param lst_data : List data adding to page
        """
        # ログ
        log.debug(self, args=None, kwargs=constant.DEBUG_MODE)
        try:
            packet = io.BytesIO()
            # Create a new PDF with Report lab
            can_page = canvas.Canvas(packet, pagesize=letter)
            # Font definition
            pdfmetrics.registerFont(TTFont(constant.FONT_NAME, constant.FONT_PATH))
            can_page.setFont(constant.FONT_NAME, constant.FONT_SIZE_DEFAULT)
            # add the "watermark" (which is the new pdf) on the existing page
            page = template.getPage(page_number - 1)
            # write Label in debug mode
            self.write_label(can_page, page)
            if not lst_position or not lst_data:
                pass
            else:
                # create list mix position and data
                lst_mix = list(zip(lst_position, lst_data))
                # add information to the specified position
                for values in lst_mix:
                    if values[0]["TYPE"] == "DATA":
                        self.write_data_pdf(can_page, values[1], values[0].get("POSITION")[0][0],
                                            values[0].get("POSITION")[0][1],
                                            values[0].get("LEN_TEXT", constant.LEN_TEXT_DEFAULT),
                                            values[0].get("CHAR_SPACE", constant.CHAR_SPACE_DEFAULT),
                                            values[0].get("FONT_SIZE", constant.FONT_SIZE_DEFAULT),
                                            values[0].get("ALIGNMENT", constant.ALIGNMENT_DEFAULT),
                                            values[0].get("COLOR", constant.COLOR_DEFAULT),
                                            values[0].get("BORDER", constant.BORDER_DEFAULT),
                                            values[0].get("BORDER_COLOR", constant.BORDER_COLOR_DEFAULT),
                                            values[0].get("BORDER_THICKNESS", constant.BORDER_THICKNESS_DEFAULT),
                                            values[0].get("BORDER_STYLE", constant.BORDER_STYLE_DEFAULT))
                    elif values[0]["TYPE"] == "CIRCLE":
                        circles_dict = dict(zip(values[0].get("VALUES"), values[0].get("POSITION")))
                        self.draw_circles(can_page, values[1], circles_dict, values[0].get("OPTION"))
                    elif values[0]["TYPE"] == "LINE":
                        xy_end = [values[0].get("POSITION")[0][0] + self.get_width_text(values[1], constant.FONT_PATH),
                                  values[0].get("POSITION")[0][1]]
                        self.draw_line(can_page, values[0].get("POSITION")[0], xy_end, values[0].get("OPTION"))
                    elif values[0]["TYPE"] == "BOX":
                        self.draw_rectangle(can_page, values[0].get("POSITION")[0][0],
                                            values[0].get("POSITION")[0][1], values[0].get("DIMENSION")[0][0],
                                            values[0].get("DIMENSION")[0][1],
                                            values[0].get("LINE_WIDTH", constant.LINE_WIDTH_DEFAULT),
                                            values[0].get("STROKE_COLOR", constant.STROKE_COLOR_DEFAULT),
                                            values[0].get("FILL_COLOR"),
                                            values[0].get("BORDER_DASH", constant.BORDER_DASH_DEFAULT),
                                            values[0].get("MODE_STROKE", constant.MODE_STROKE_DEFAULT),
                                            values[0].get("MODE_FILL", constant.MODE_FILL_DEFAULT),
                                            values[0].get("TEXT_COLOR", constant.TEXT_COLOR_DEFAULT))

            can_page.save()
            # move to the beginning of the StringIO buffer
            packet.seek(0)
            # Initialize PdfFileReader
            new_pdf = PdfFileReader(packet)
            # merge information to page
            # rotate angle is 0 degree
            if page['/Rotate'] == 0:
                page.mergePage(new_pdf.getPage(0))
            # rotate angle is 90 degree
            elif page['/Rotate'] == 90:
                page.mergeRotatedScaledTranslatedPage(new_pdf.getPage(0), page['/Rotate'], 1, 595, 0, expand=False)
            # rotate angle is 180 degree
            elif page['/Rotate'] == 180:
                page.mergeRotatedScaledTranslatedPage(new_pdf.getPage(0), page['/Rotate'], 1, 845, 595, expand=False)
            # rotate angle is 270 degree
            else:
                page.mergeRotatedScaledTranslatedPage(new_pdf.getPage(0), page['/Rotate'], 1, 0, 595, expand=False)
            # start rotate pdf file
            # get layout from .json file
            layout = self.json_data["CONFIG"].get("layout")
            # file direction is portrait and file direction on default mode
            if layout == "portrait" or layout is None:
                if page.mediaBox.upperRight[0] < page.mediaBox.upperRight[1]:
                    if page['/Rotate'] == 90:
                        page.rotateClockwise(270)
                elif page.mediaBox.upperRight[0] > page.mediaBox.upperRight[1]:
                    if page['/Rotate'] != 270:
                        page.rotateClockwise(270)
            # file direction is landscape
            elif layout == "landscape":
                if page.mediaBox.upperRight[0] > page.mediaBox.upperRight[1]:
                    if page['/Rotate'] == 270:
                        page.rotateClockwise(270)
                elif page.mediaBox.upperRight[0] < page.mediaBox.upperRight[1]:
                    if page['/Rotate'] != 90:
                        page.rotateClockwise(270)
            # file direction is not change
            elif layout == "free":
                pass
            # end rotate file pdf
            log.info(self)
            return page
        except:
            # 例外処理
            log.error(traceback.format_exc())

    def get_width_text(self, text, font_path, f_size=constant.FONT_SIZE_DEFAULT):
        """
        Calculate the length of a string insert into PDF document
        :param f_size: font size default = 12
        :param text: String insert into PDF document
        :param font_path: Path of font
        :return: Width of text
        """
        # ログ
        log.debug(self)
        try:
            # Set font of text
            font = ImageFont.truetype(font_path, f_size)
            # Get size of text
            size = font.getsize(str(text))
            log.info(self)
            # Width of text
            return size[0]
        except:
            # 例外処理
            log.error(traceback.format_exc())

    def get_height_text(self, text, font_path, f_size=constant.FONT_SIZE_DEFAULT):
        """
        Calculate the length of a string insert into PDF document
        :param f_size: font size default = 12
        :param text: String insert into PDF document
        :param font_path: Path of font
        :return: Width of text
        """
        # ログ
        log.debug(self)
        try:
            # Set font of text
            font = ImageFont.truetype(font_path, f_size)
            # Get size of text
            size = font.getsize(str(text))
            log.info(self)
            # Width of text
            return size[1]
        except:
            # 例外処理
            log.error(traceback.format_exc())

    def write_data_pdf(self, can_page, data, cx_start, cy_start, cx_end=100, char_spacing=constant.CHAR_SPACING,
                        f_size=constant.FONT_SIZE_DEFAULT,
                        alignment="left", color=constant.COLOR_DEFAULT, border=constant.BORDER_DEFAULT,
                        border_color=constant.BORDER_COLOR_DEFAULT, border_thickness=constant.BORDER_THICKNESS_DEFAULT,
                        border_style=constant.BORDER_STYLE_DEFAULT):
        """
        Check Over Width Text of data then add watermark
        :param data: data to write into pdf page
        :param can_page: canvas object
        :param cx_start: positions start to write data on
        :param cy_start: positions end to write data on
        :param cx_end: positions end to change status default = 100
        :param char_spacing: values to set char spacing
        :param f_size: font size setting default = 10
        :param alignment: alignment to write data on
        :param color : color text
        :param border : border text
        :param border_color : color border text
        :param border_thickness : thichkness border
        :param border_style : border styles
        :return: None
        """
        # ログ
        log.debug(self, args=None, kwargs=constant.DEBUG_MODE)
        try:
            # get width of data to write into pdf page
            tmp = self.get_width_text(data, constant.FONT_PATH, f_size)
            # Set font size smaller
            if tmp > cx_end:
                f_size = f_size - 4
            # Get text width under font size
            gWidth = self.get_width_text(data, constant.FONT_PATH, f_size)
            # Get text height under font size
            gHeight = self.get_height_text(data, constant.FONT_PATH, f_size)
            can_page.setFont(constant.FONT_NAME, f_size)
            self.style_data_pdf(can_page, data, cx_start, cy_start, gWidth, gHeight, f_size, char_spacing, color, border,
                                border_color, border_thickness, border_style)
            # write data to pdf file
            if alignment == constant.ALIGN_CENTER:
                # center alignment
                can_page.drawCentredString(cx_start, cy_start, str(data), None, char_spacing)
            elif alignment == constant.ALIGN_RIGHT:
                # Right alignment
                can_page.drawRightString(cx_start, cy_start, str(data), None, char_spacing)
            else:
                # Left alignment
                can_page.drawString(cx_start, cy_start, str(data), None, char_spacing)
        except:
            # 例外処理
            # log.error(traceback.format_exc())
            constant.get_error(constant.ERROR_003)

    def style_data_pdf(self, can_page, data, cx_start, cy_start, tmp, hmp, f_size, char_spacing=constant.CHAR_SPACING, color=constant.COLOR_DEFAULT,
                        border=constant.BORDER_DEFAULT,
                        border_color=constant.BORDER_COLOR_DEFAULT, border_thickness=constant.BORDER_THICKNESS_DEFAULT,
                        border_style=constant.BORDER_STYLE_DEFAULT):
        """
        Style data PDF
        :param data: data to write into pdf page
        :param can_page: canvas object
        :param cx_start: positions start to write data on
        :param cy_start: positions end to write data on
        :param f_size: font size setting default = 10
        :param tmp : with text
        :param hmp : height text
        :param color : color text
        :param border : border text
        :param border_color : color border text
        :param border_thickness : thichkness border
        :param border_style : border styles
        """
        try:
            if border_style == "dash":
                # Border simple
                can_page.setDash(3, 1)
            elif border_style == "dots":
                # Border dots
                can_page.setDash(1, 1)
            elif border_style == "complex":
                # Border complex
                can_page.setDash([1, 1, 3, 3, 1, 4, 4, 1], 0)
            # Thickness border
            can_page.setLineWidth(border_thickness)
            # Color border
            can_page.setStrokeColor(border_color)
            # Color text
            can_page.setFillColor(color)
            x1 = cx_start - 5
            y1 = cy_start - 5
            x2 = tmp + tmp * 0.16 + cx_start
            y2 = hmp + cy_start
            # detect the language of the text
            data_str = str(data)
            language = langid.classify(data_str)
            len_data = len(data) -1
            rex = 0.16
            if data_str.isdigit() and f_size < 5:
                rex = 1
            elif data_str.isdigit() and f_size < 8:
                rex = 0.5
            elif data_str.isdigit() and f_size < 12:
                rex = 0.3
            elif data_str.isdigit() and f_size < 17:
                rex = 0.19
            elif data_str.isdigit() and f_size < 30:
                rex = 0.18
            elif f_size <= 6:
                if language[0] == "en":
                    rex = 0.6
                else:
                    rex = 0.4
            elif f_size <= 10:
                if language[0] == "en":
                    rex = 0.45
                else:
                    rex = 0.23
            elif f_size <= 16:
                if language[0] == "en":
                    rex = 0.3
                else:
                    rex = 0.15
            elif f_size <= 30:
                if language[0] == "en":
                    rex = 0.18
                else:
                    rex = 0.1
            else:
                rex = 0.08
            if char_spacing == 1.7:
                x2 = tmp + tmp * rex + cx_start
            else:
                if len(data) >= 30:
                    len_data = len(data) -10
                elif len(data) >= 20:
                    if char_spacing < 10:
                        len_data = len(data) -4
                    elif char_spacing < 20:
                        len_data = len(data) -3
                    else:
                        len_data = len(data) -2
                elif len(data) >= 10:
                    if char_spacing < 10:
                        len_data = len(data) -3
                    elif char_spacing < 20:
                        len_data = len(data) -2
                    else:
                        len_data = len(data) -1
                else:
                    len_data = len(data) -1
                x2 = tmp + tmp * rex + cx_start + char_spacing * len_data
            if border:
                if type(border) == list:
                    bd_df_left = (0, 0, 0, 0)
                    bd_df_bottom = (0, 0, 0, 0)
                    bd_df_top = (0, 0, 0, 0)
                    bd_df_right = (0, 0, 0, 0)
                    for i in border:
                        if (i == "left"):
                            # Border left
                            bd_df_left = (x1, y1, x1, y2)
                        if (i == "bottom"):
                            # Border bottom
                            bd_df_bottom = (x1, y1, x2, y1)
                        if (i == "top"):
                            # Border top
                            bd_df_top = (x1, y2, x2, y2)
                        if (i == "right"):
                            # Border right
                            bd_df_right = (x2, y1, x2, y2)
                        can_page.lines([bd_df_left, bd_df_bottom, bd_df_top, bd_df_right])
                else:
                    # Border full
                    can_page.lines([(x1, y1, x1, y2), (x1, y1, x2, y1), (x1, y2, x2, y2), (x2, y1, x2, y2)])
        except:
            # 例外処理
            # log.error(traceback.format_exc())
            constant.get_error(constant.ERROR_003)

    def write_label(self, can_page, page):
        """
        Write number of function to top right in DEBUG MODE
        :param can_page: canvas object
        :return: None
        """
        log.debug(self, args=None, kwargs=constant.DEBUG_MODE)
        try:
            if constant.DEBUG_MODE:
                log.debug(self, args=None, kwargs=constant.DEBUG_MODE)
                can_page.setFont(constant.FONT_NAME, constant.FONT_SIZE_DEFAULT)
                # rotate angle is 0 degree
                if page['/Rotate'] == 0:
                    if page.mediaBox.upperRight[0] > page.mediaBox.upperRight[1]:
                        can_page.drawString(constant.LABEL_DEFAULT_POSITION["PDF_ROTATE"][0],
                                        constant.LABEL_DEFAULT_POSITION["PDF_ROTATE"][1], self.kinoubango, None,
                                        constant.CHAR_SPACING)
                    else:
                        can_page.drawString(constant.LABEL_DEFAULT_POSITION["PDF"][0],
                                        constant.LABEL_DEFAULT_POSITION["PDF"][1], self.kinoubango, None,
                                        constant.CHAR_SPACING)
                # rotate angle is 90 degree or rotate angle is 180 degree
                elif page['/Rotate'] == 90 or page['/Rotate'] == 180:
                    can_page.drawString(constant.LABEL_DEFAULT_POSITION["PDF_ROTATE"][0],
                                    constant.LABEL_DEFAULT_POSITION["PDF_ROTATE"][1], self.kinoubango, None,
                                    constant.CHAR_SPACING)
                # rotate angle is 270 degree
                else:
                    can_page.drawString(constant.LABEL_DEFAULT_POSITION["PDF"][0],
                                    constant.LABEL_DEFAULT_POSITION["PDF"][1], self.kinoubango, None,
                                    constant.CHAR_SPACING)
            else:
                can_page.drawString(constant.LABEL_DEFAULT_POSITION["PDF"][0],
                                    constant.LABEL_DEFAULT_POSITION["PDF"][1], "", None,
                                    constant.CHAR_SPACING)
        except:
            # 例外処理
            # log.error(traceback.format_exc())
            constant.get_error(constant.ERROR_003)

    def draw_circles(self, can_page, key, circles_dict=dict, circle_options=list):
        """
        Drawing an oval to choose an option on file PDF
        :param can_page: canvas object
        :param key: values of position need to draw circles
        :param circles_dict: circles dictionary
        :param circle_options: List option height, width
        :return:
        """
        # ログ
        log.debug(self)
        try:
            # Get position by key
            pos = circles_dict.get(key)
            if not pos:
                log.error("************* WARNING *************")
                log.error("============= Please check data on file csv values : %s =============" % key)
            else:
                # draw circle
                can_page.ellipse(pos[0], pos[1], pos[0] + circle_options[0],
                                pos[1] + circle_options[1], stroke=1, fill=0)
        except:
            # 例外処理
            # log.error(traceback.format_exc())
            constant.get_error(constant.ERROR_003)

    def draw_line(self, can_page, xy_start, xy_end, line_options=constant.DEFAULT_LINE_OPTIONS):
        """
        Drawing the line
        :param can_page: canvas object
        :param xy_start: positions start to draw line on
        :param xy_end: positions end to draw line on
        :param line_options: Line formatting (SOLID or DASH)
        :return:
        """
        log.debug(self)
        try:
            if line_options.get("LINE_TYPE") == "DASH":
                dash_option = None
                if line_options.get("DASH"):
                    dash_option = line_options.get("DASH")
                can_page.setDash(dash_option)
            can_page.line(xy_start[0], xy_start[1], xy_end[0], xy_end[1])
            # switch from dashed to solid
            can_page.setDash(())
        except:
            # 例外処理
            # log.error(traceback.format_exc())
            constant.get_error(constant.ERROR_003)

    def draw_rectangle(self, can_page, x_start, y_start, width_rect, height_rect, line_width, stroke_color, fill_color,
                        dash_style, stroke_mode, fill_mode, text_color):
        """
        Drawing an rectangle to choose an option on file PDF.
        Mode write text in rectangle is set in constant.py file
        :param can_page: canvas object
        :param x_start: values of position need to draw circles
        :param y_start: values of position need to draw circles
        :param width_rect: width of rectangle
        :param height_rect: height of rectangle
        :param line_width: width of rectangle borders
        :param stroke_color: color rectangle border
        :param fill_color: color background rectangle
        :param dash_style: style of border
        :param stroke_mode: mode draw rectangle border
        :param fill_mode: mode fill color background rectangle
        :param text_color: color of text
        :return:
        """
        # ログ
        log.debug(self)
        try:
            can_page.setLineWidth(line_width)
            can_page.setStrokeColor(stroke_color)
            if fill_color is None:
                fill_mode = 0
            else:
                can_page.setFillColor(fill_color)
            can_page.setDash(dash_style)
            can_page.rect(x_start, y_start, width_rect, height_rect, stroke=stroke_mode, fill=fill_mode)
            can_page.setFillColor(text_color, alpha=None)
        except:
            # 例外処理
            # log.error(traceback.format_exc())
            constant.get_error(constant.ERROR_003)


class ExcelFilesMaker(CommonFilesMaker):
    """
    Class printing Excel object
    """

    def __init__(self, module, category, kinoubango, activity_code, config):
        super().__init__(module, category, kinoubango, activity_code, config)

    # open workbook with file excel.xls
    def init_workbook(self, template):
        # ログ
        log.debug(self, args=None, kwargs=constant.DEBUG_MODE)
        try:
            if template.endswith("xls"):
                # Initial xlrd.Book object
                wb_template = open_workbook(template, formatting_info=True, encoding_override="cp1251")
                return wb_template
            elif template.endswith("xlsx"):
                wb_template = load_workbook(template)
                return wb_template
        except:
            # 例外処理
            # log.error(traceback.format_exc())
            constant.get_error(constant.ERROR_002)

    # create workbook with file excel.xls
    def create_workbook_ZU_P99(self, template, sheet_number, wb_output, lst_position, lst_data):

        # ログ
        log.debug(self, args=None, kwargs=constant.DEBUG_MODE)
        try:
            if not lst_position or not lst_data:
                # Debug mode
                if constant.DEBUG_MODE:
                    # Log information
                    log.debug(self, args=None, kwargs=constant.DEBUG_MODE)
                    # .xls file
                    if template.endswith("xls"):
                        # Write label to .xls file
                        self.write_labels_xls(wb_output, sheet_number)
                    # .xlsx file
                    elif template.endswith("xlsx"):
                        # Write label to .xlsx file
                        self.write_labels_xlsx(wb_output, sheet_number)
            elif template.endswith("xls"):
                # create list mix position and data
                lst_mix = list(zip(lst_position, lst_data))
                # write Label in debug mode
                if constant.DEBUG_MODE:
                    log.debug(self, args=None, kwargs=constant.DEBUG_MODE)
                    self.write_labels_xls(wb_output, sheet_number)
                # Write all data into 1 worksheet
                for values in lst_mix:
                    if values[0]["TYPE"] == "DATA":
                        self.write_data_xls(wb_output, sheet_number, values[0]["POSITION"][0][0],
                                            values[0]["POSITION"][0][1], values[1], values[0])
                    elif values[0]["TYPE"] == "LINE":
                        self.write_border_xls(wb_output, sheet_number, values[0].get('POSITION'),
                                                values[0].get('OPTION'))
            elif template.endswith("xlsx"):
                # create list mix position and data
                lst_mix = list(zip(lst_position, lst_data))
                # write Label in debug mode
                if constant.DEBUG_MODE:
                    log.debug(self, args=None, kwargs=constant.DEBUG_MODE)
                    self.write_labels_xlsx(wb_output, sheet_number)
                # Write all data into 1 worksheet
                for values in lst_mix:
                    if values[0]["TYPE"] == "DATA":
                        self.write_data_xlsx(wb_output, sheet_number, values[0]["POSITION"][0][0],
                                                values[0]["POSITION"][0][1], values[1], values[0])
                    elif values[0]["TYPE"] == "LINE":
                        self.write_border_xlsx(wb_output, sheet_number, values[0].get('POSITION'),
                                                values[0].get('OPTION'))
        except:
            # 例外処理
            # log.error(traceback.format_exc())
            constant.get_error(constant.ERROR_003)

    # write data and set style with file excel.xls
    @staticmethod
    def write_data_xls(wb_output, sheet_number, row, col, data, style_att=dict):
        try:
            if not style_att:
                wb_output.get_sheet(sheet_number).write(row, col, data)
            else:
                # Setting for formatting cell.
                fmt_cell = xlwt.Style.easyxf(style_att.get("STYLE", ""), style_att.get("FORMAT_NUMBER", None),
                                                style_att.get("FIELD_SEP", ","),
                                                style_att.get("LINE_SEP", ";"), style_att.get("INTRO_SEP", ":"),
                                                style_att.get("ESC_CHAR", "\\"),
                                                style_att.get("DEBUG", False))
                wb_output.get_sheet(sheet_number).write(row, col, data, fmt_cell)
        except:
            constant.get_error(constant.ERROR_003)

    # write data and set style with file excel.xlsx
    @staticmethod
    def write_data_xlsx(wb_template, sheet_number, row, col, data, style_att=dict):
        try:
            font_format = ExcelFilesMaker.convert_font_attribute(style_att.get("FONT", constant.FONT_SETTING))

            pattern_format = ExcelFilesMaker.convert_pattern_fill_attribute(
                style_att.get("PATTERN_FILL", constant.PATTERN_FILL_SETTING))

            border_format = ExcelFilesMaker.convert_border_attribute(style_att.get("BORDER", constant.BORDER_SETTING))

            alignment_format = ExcelFilesMaker.convert_alignment_attribute(
                style_att.get("ALIGNMENT", constant.ALIGNMENT_SETTING))

            num_format = ExcelFilesMaker.convert_number_format_attribute(
                style_att.get("NUMBER_FORMAT", constant.NUMBER_FORMAT_SETTING))

            protection_format = ExcelFilesMaker.convert_protection_attribute(
                style_att.get("PROTECTION", constant.PROTECTION_SETTING))

            wb_template[sheet_number].cell(row=row + 1, column=col + 1).value = data
            wb_template[sheet_number].cell(row=row + 1, column=col + 1).font = font_format
            wb_template[sheet_number].cell(row=row + 1, column=col + 1).fill = pattern_format
            wb_template[sheet_number].cell(row=row + 1, column=col + 1).number_format = num_format
            wb_template[sheet_number].cell(row=row + 1, column=col + 1).protection = protection_format
            wb_template[sheet_number].cell(row=row + 1, column=col + 1).alignment = alignment_format
            wb_template[sheet_number].cell(row=row + 1, column=col + 1).border = border_format
        except:
            constant.get_error(constant.ERROR_003)

    def write_labels_xls(self, wb_output, sheet_number, row=constant.LABEL_DEFAULT_POSITION["EXCEL"][0],
                            col=constant.LABEL_DEFAULT_POSITION["EXCEL"][1]):
        try:
            wb_output.get_sheet(sheet_number).write(row, col, self.kinoubango)
        except:
            constant.get_error(constant.ERROR_003)

    def write_labels_xlsx(self, wb_template, sheet_number, row=constant.LABEL_DEFAULT_POSITION["EXCEL"][0],
                            col=constant.LABEL_DEFAULT_POSITION["EXCEL"][1]):
        try:
            wb_template[sheet_number].cell(row=row + 1, column=col + 1).value = self.kinoubango
        except:
            constant.get_error(constant.ERROR_003)

    def on_write_border_xls(self, keys, param, style, pos, wb_output, s_number):
        """

        """
        try:
            borders = xlwt.Borders()
            for key in keys:
                setattr(borders, key, param["bd_style"])
                setattr(borders, key + '_colour', param["bd_color"])

            style.borders = borders
            wb_output.get_sheet(s_number).write(pos["row"], pos["col"], style=style)
        except:
            constant.get_error(constant.ERROR_003)

    def write_border_xls(self, wb_output, sheet_number, cell_ranges, border_options):
        """
        Apply styles to a range of cells as if they were a single cell.
        :param wb_output:  Excel worksheet instance
        :param sheet_number: Name of sheet
        :param cell_ranges: An excel cell range to style
        :param border_options: An openpyxl Border
        Border style: THIN, MEDIUM, DASHED, DOTTED, THICK, DOUBLE, HAIR
        """
        # ログ
        log.debug(self)
        try:
            col1 = cell_ranges[0][0] - 1
            row1 = cell_ranges[0][1] - 1
            col2 = cell_ranges[1][0] - 1
            row2 = cell_ranges[1][1] - 1
            # Border styles
            bd_style_dicts = {
                'thin': xlwt.Borders.THIN,
                'medium': xlwt.Borders.MEDIUM,
                'dashed': xlwt.Borders.DASHED,
                'dotted': xlwt.Borders.DOTTED,
                'thick': xlwt.Borders.THICK,
                'double': xlwt.Borders.DOUBLE,
                'hair': xlwt.Borders.HAIR,
            }
            bd_style = bd_style_dicts.get('thin')
            bd_style_name = str(border_options.get('style')).lower()
            if bd_style_dicts.get(bd_style_name):
                bd_style = bd_style_dicts.get(bd_style_name)
            # Border color
            bd_color_dicts = ['black', 'blue', 'brown', 'gold', 'green', 'red', 'pink', 'white', 'yellow']
            n_color = 'black'
            if border_options.get('color') and border_options.get('color').lower() in bd_color_dicts:
                n_color = border_options.get('color').lower()
            bd_color = xlwt.Style.colour_map[n_color]
            bd_style_option = {
                'bd_style': bd_style,
                'bd_color': bd_color
            }
            style = xlwt.XFStyle()
            # draw border left + right
            for j in range(row1, row2 + 1):
                pos = {"row": j, "col": 0}
                if "left" in border_options.get('border'):
                    key = ["left"] if (col1 == 0) else ["right"]
                    pos["col"] = col1 if (col1 == 0) else col1 - 1
                    self.on_write_border_xls(key, bd_style_option, style, pos, wb_output, sheet_number)
                if "right" in border_options.get('border'):
                    pos["col"] = col2 + 1
                    self.on_write_border_xls(["left"], bd_style_option, style, pos, wb_output, sheet_number)
            # draw border bottom + top
            for i in range(col1, col2 + 1):
                pos = {"row": 0, "col": i}
                if "bottom" in border_options.get('border'):
                    key = ["top"]
                    pos["row"] = row2 + 1
                    self.on_write_border_xls(key, bd_style_option, style, pos, wb_output, sheet_number)
                if "top" in border_options.get('border'):
                    if row1 == 0:
                        key = ["top", "left"] if (i == 0) else ["top"]
                        pos["row"] = row1
                    else:
                        key = ["top", "left"] if (i == 0) else ["bottom"]
                        pos["row"] = row1 if (i == 0) else row1 - 1
                    self.on_write_border_xls(key, bd_style_option, style, pos, wb_output, sheet_number)
        except:
            # 例外処理
            # log.error(traceback.format_exc())
            constant.get_error(constant.ERROR_003)

    def write_border_xlsx(self, wb_template, sheet_number, cell_range, border_options):
        """
        Apply styles to a range of cells as if they were a single cell.
        :param wb_template:  Excel worksheet instance
        :param sheet_number: Name of sheet
        :param cell_range: An excel cell range to style
        :param border_options: An openpyxl Border
        Border style: double, mediumDashed, hair, thick, mediumDashDotDot, mediumDashDot, dashed, dotted, dashDot, dashDotDot, medium,
        slantDashDot, thin,
        """
        # ログ
        log.debug(self)
        try:
            cell_1 = xl_rowcol_to_cell(cell_range[0][0], cell_range[0][1])
            cell_2 = xl_rowcol_to_cell(cell_range[1][0], cell_range[1][1])
            join_cell_range = ':'.join([cell_1, cell_2])
            rows = wb_template[sheet_number][join_cell_range]
            # Border color
            bd_color_dicts = {
                'black': "000000",
                'blue': "0000ff",
                'brown': "a52a2a",
                'gold': "ffd700",
                'green': "008000",
                'red': "ff0000",
                'pink': "ffc0cb",
                'white': "ffffff",
                'yellow': "ffff00"
            }
            bd_color = 'black'
            if border_options.get('color'):
                bd_color = border_options.get('color').lower()
            color = bd_color_dicts.get(bd_color)
            # Border style
            bd_style_dicts = ["double", "mediumDashed", "hair", "thick", "mediumDashDotDot", "mediumDashDot", "dashed",
                                "dotted", "dashDot", "dashDotDot", "medium", "slantDashDot", "thin"]
            style = "thin"
            if border_options.get('style') and border_options.get('style').lower() in bd_style_dicts:
                style = border_options.get('style').lower()

            bd_style = Side(border_style=style, color=color)
            border = Border(top=bd_style, left=bd_style, right=bd_style, bottom=bd_style)
            left = ""
            right = ""
            for bd in border_options.get('border'):
                if bd == 'left':
                    left = Border(left=border.left)
                elif bd == 'top':
                    top = Border(top=border.top)
                    for cell in rows[0]:
                        cell.border = cell.border + top
                elif bd == 'right':
                    right = Border(right=border.right)
                else:
                    bottom = Border(bottom=border.bottom)
                    for cell in rows[-1]:
                        cell.border = cell.border + bottom

            for row in rows:
                left_ = row[0]
                right_ = row[-1]
                if 'left' in border_options.get('border'):
                    left_.border = left_.border + left
                if 'right' in border_options.get('border'):
                    right_.border = right_.border + right
        except:
            # 例外処理
            # log.error(traceback.format_exc())
            constant.get_error(constant.ERROR_003)

    @staticmethod
    def convert_font_attribute(font_dic=dict):
        return Font(
            name=font_dic.get("name", "Calibri"),
            size=font_dic.get("size", 11),
            bold=font_dic.get("bold", False),
            italic=font_dic.get("italic", False),
            vertAlign=font_dic.get("vertAlign", None),
            underline=font_dic.get("underline", None),
            strike=font_dic.get("strike", False),
            color=font_dic.get("color", "FF000000"),
        )

    @staticmethod
    def convert_pattern_fill_attribute(pattern_dic=dict):
        return PatternFill(
            fill_type=pattern_dic.get("fill_type", None),
            start_color=pattern_dic.get("start_color", "FFFFFFFF"),
            end_color=pattern_dic.get("end_color", 'FF000000'),
        )

    @staticmethod
    def convert_border_attribute(border_dic=dict):
        left_border_styles = None
        left_colors = "FF000000"
        if border_dic.get("left", None):
            left_border_styles = border_dic.get("left").get("border_style", None)
            left_colors = border_dic.get("left").get("color", "FF000000")

        right_border_styles = None
        right_colors = "FF000000"
        if border_dic.get("right", None):
            right_border_styles = border_dic.get("right").get("border_style", None)
            right_colors = border_dic.get("right").get("color", "FF000000")

        top_border_styles = None
        top_colors = "FF000000"
        if border_dic.get("top", None):
            top_border_styles = border_dic.get("top").get("border_style", None)
            top_colors = border_dic.get("top").get("color", "FF000000")

        bottom_border_styles = None
        bottom_colors = "FF000000"
        if border_dic.get("bottom", None):
            bottom_border_styles = border_dic.get("bottom").get("border_style", None)
            bottom_colors = border_dic.get("bottom").get("color", "FF000000")

        diagonal_border_styles = None
        diagonal_colors = "FF000000"
        if border_dic.get("diagonal", None):
            diagonal_border_styles = border_dic.get("diagonal").get("border_style", None)
            diagonal_colors = border_dic.get("diagonal").get("color", "FF000000")

        outline_border_styles = None
        outline_colors = "FF000000"
        if border_dic.get("outline", None):
            outline_border_styles = border_dic.get("outline").get("border_style", None)
            outline_colors = border_dic.get("outline").get("color", "FF000000")

        vertical_border_styles = None
        vertical_colors = "FF000000"
        if border_dic.get("vertical", None):
            vertical_border_styles = border_dic.get("vertical").get("border_style", None)
            vertical_colors = border_dic.get("vertical").get("color", "FF000000")

        horizontal_border_styles = None
        horizontal_colors = "FF000000"
        if border_dic.get("horizontal", None):
            horizontal_border_styles = border_dic.get("horizontal").get("border_style", None)
            horizontal_colors = border_dic.get("horizontal").get("color", "FF000000")

        return Border(
            left=Side(border_style=left_border_styles, color=left_colors),
            right=Side(border_style=right_border_styles, color=right_colors),
            top=Side(border_style=top_border_styles, color=top_colors),
            bottom=Side(border_style=bottom_border_styles, color=bottom_colors),
            diagonal=Side(border_style=diagonal_border_styles, color=diagonal_colors),
            diagonal_direction=border_dic.get("diagonal_direction", 0),
            outline=Side(border_style=outline_border_styles, color=outline_colors),
            vertical=Side(border_style=vertical_border_styles, color=vertical_colors),
            horizontal=Side(border_style=horizontal_border_styles, color=horizontal_colors),
        )

    @staticmethod
    def convert_alignment_attribute(alignment_dic=dict):
        return Alignment(
            horizontal=alignment_dic.get("horizontal", "general"),
            vertical=alignment_dic.get("vertical", "bottom"),
            text_rotation=alignment_dic.get("text_rotation", 0),
            wrap_text=alignment_dic.get("wrap_text", False),
            shrink_to_fit=alignment_dic.get("shrink_to_fit", False),
            indent=alignment_dic.get("indent", 0),
        )

    @staticmethod
    def convert_number_format_attribute(number_format_dic=dict):
        return number_format_dic.get("number_format", "General")

    @staticmethod
    def convert_protection_attribute(protection_dic=dict):
        return Protection(
            locked=protection_dic.get("locked", True),
            hidden=protection_dic.get("hidden", False),
        )


class FilesMaker(PDFFilesMaker, ExcelFilesMaker):
    """
    
    """
    def __init__(self, module, category, kinoubango, activity_code, config):
        super().__init__(module, category, kinoubango, activity_code, config)

# ------------------------------------------------------------
